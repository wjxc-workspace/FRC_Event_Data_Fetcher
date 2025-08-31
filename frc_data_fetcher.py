#!/usr/bin/env python3
"""
FRC Event Data Fetcher
Fetches team statistics, EPA rankings, and awards from The Blue Alliance and Statbotics APIs.
"""

import os
import sys
import logging
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor, as_completed

import tbaapiv3client
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from statbotics import Statbotics

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()


@dataclass
class Config:
    """Configuration settings for API connections"""
    first_api_host: str = 'https://frc-api.firstinspires.org/v3.0/'
    first_api_username: str = os.getenv('FIRST_API_USERNAME', '')
    first_api_auth_token: str = os.getenv('FIRST_API_AUTH_TOKEN', '')
    tba_api_host: str = 'https://www.thebluealliance.com/api/v3'
    tba_api_key: str = os.getenv('TBA_API_KEY', '')
    
    def validate(self) -> None:
        """Validate required API keys are present"""
        if not self.tba_api_key:
            raise ValueError("TBA_API_KEY not found in environment variables")


@dataclass
class TeamStats:
    """Data structure for team statistics"""
    epa: float | str
    rank: int | str
    
    @classmethod
    def empty(cls) -> 'TeamStats':
        """Create empty stats object for error cases"""
        return cls(epa='N/A', rank='N/A')


class FRCDataFetcher:
    """Main class for fetching FRC team data"""
    
    def __init__(self, config: Config):
        self.config = config
        self.config.validate()
        
        # Initialize TBA API client
        tba_config = tbaapiv3client.Configuration(
            host=self.config.tba_api_host,
            api_key={'X-TBA-Auth-Key': self.config.tba_api_key}
        )
        self.tba_client = tbaapiv3client.ApiClient(tba_config)
        self.tba_event_api = tbaapiv3client.EventApi(self.tba_client)
        self.tba_team_api = tbaapiv3client.TeamApi(self.tba_client)
        
        # Initialize Statbotics client
        self.sb = Statbotics()
        
        # Cache for API responses
        self._cache: Dict[str, Any] = {}
    
    def get_event_teams(self, event_key: str) -> List[int]:
        """
        Fetch list of teams participating in an event
        
        Args:
            event_key: The event key (e.g., '2024txhou')
            
        Returns:
            Sorted list of team numbers
        """
        cache_key = f"teams_{event_key}"
        if cache_key in self._cache:
            return self._cache[cache_key]
        
        try:
            response = self.tba_event_api.get_event_teams_simple(
                event_key,
                if_modified_since='Cache-Control'
            )
            teams = sorted([team.team_number for team in response])
            self._cache[cache_key] = teams
            return teams
        except Exception as e:
            logger.error(f"Failed to fetch teams for event {event_key}: {e}")
            raise RuntimeError(f"Could not fetch teams for event {event_key}. Please check your internet connection and API key.")
    
    def get_team_statbotics(self, team_number: int, year: int) -> TeamStats:
        """
        Fetch Statbotics EPA data for a team in a specific year
        
        Args:
            team_number: FRC team number
            year: Competition year
            
        Returns:
            TeamStats object with EPA and rank data
        """
        cache_key = f"sb_{team_number}_{year}"
        if cache_key in self._cache:
            return self._cache[cache_key]
        
        try:
            response = self.sb.get_team_year(team_number, year, ['epa'])
            stats = TeamStats(
                epa=round(response['epa']['total_points']['mean'], 2),
                rank=response['epa']['ranks']['total']['rank']
            )
            self._cache[cache_key] = stats
            return stats
        except Exception as e:
            logger.debug(f"No Statbotics data for team {team_number} in {year}: {e}")
            return TeamStats.empty()
    
    def get_team_events(self, team_number: int, year: int) -> List[str]:
        """
        Get list of events a team participated in during a year
        
        Args:
            team_number: FRC team number
            year: Competition year
            
        Returns:
            List of event keys
        """
        cache_key = f"events_{team_number}_{year}"
        if cache_key in self._cache:
            return self._cache[cache_key]
        
        try:
            events = self.tba_team_api.get_team_events_by_year_keys(
                f'frc{team_number}', 
                year
            )
            self._cache[cache_key] = events
            return events
        except Exception as e:
            logger.debug(f"No events found for team {team_number} in {year}: {e}")
            return []
    
    def get_team_event_awards(self, team_number: int, event_key: str) -> List[str]:
        """
        Get awards won by a team at a specific event
        
        Args:
            team_number: FRC team number
            event_key: Event key
            
        Returns:
            List of award descriptions
        """
        cache_key = f"awards_{team_number}_{event_key}"
        if cache_key in self._cache:
            return self._cache[cache_key]
        
        try:
            response = self.tba_event_api.get_team_event_awards(
                f'frc{team_number}',
                event_key
            )
            awards = [f"{award.event_key} - {award.name}" for award in response]
            self._cache[cache_key] = awards
            return awards
        except Exception as e:
            logger.debug(f"No awards found for team {team_number} at {event_key}: {e}")
            return []
    
    def fetch_team_year_data(self, team_number: int, year: int) -> Dict[str, Any]:
        """
        Fetch all data for a team in a specific year
        
        Args:
            team_number: FRC team number
            year: Competition year
            
        Returns:
            Dictionary containing stats and awards
        """
        # Get Statbotics data
        stats = self.get_team_statbotics(team_number, year)
        
        # Get awards from all events
        events = self.get_team_events(team_number, year)
        all_awards = []
        for event in events:
            awards = self.get_team_event_awards(team_number, event)
            all_awards.extend(awards)
        
        return {
            'epa': stats.epa,
            'rank': stats.rank,
            'awards': '\n'.join(all_awards) if all_awards else ''
        }
    
    def fetch_team_data(self, team_number: int, start_year: int, end_year: int) -> List[Any]:
        """
        Fetch multiple years of data for a team
        
        Args:
            team_number: FRC team number
            start_year: First year to fetch
            end_year: Last year to fetch (inclusive)
            
        Returns:
            List of data items for Excel row
        """
        items = [team_number]
        
        for year in range(start_year, end_year + 1):
            year_data = self.fetch_team_year_data(team_number, year)
            items.extend([year_data['epa'], year_data['rank'], year_data['awards']])
        
        return items
    
    def export_to_excel(self, event_year: int, event_code: str, 
                       teams: List[int], years_to_fetch: int,
                       max_workers: int = 5) -> None:
        """
        Export all team data to Excel file with parallel processing
        
        Args:
            event_year: Year of the event
            event_code: Event code
            teams: List of team numbers
            years_to_fetch: Number of years of history to fetch
            max_workers: Maximum number of parallel threads
        """
        foldername = "output"
        filename = f"{event_year}{event_code}.xlsx"
        fullname = os.path.join(foldername, filename)
        
        # Create output folder if it doesn't exist
        if not os.path.exists(foldername):
            os.makedirs(foldername)
            logger.info(f"Created output folder: {foldername}")
        
        # Remove existing file if it exists
        if os.path.exists(fullname):
            os.remove(fullname)
            logger.info(f"Removed existing file: {filename}")
        
        # Create workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.title = f"{event_year} {event_code} Data"
        
        # Create headers
        headers = ['Team']
        start_year = event_year - years_to_fetch + 1
        for year in range(start_year, event_year + 1):
            headers.extend([f'{year} EPA', f'{year} Rank', f'{year} Awards'])
        ws.append(headers)
        
        # Adjust column widths for better readability
        ws.column_dimensions['A'].width = 10
        for i in range(1, years_to_fetch + 1):
            col_base = 1 + (i - 1) * 3
            ws.column_dimensions[chr(65 + col_base)].width = 12  # EPA
            ws.column_dimensions[chr(65 + col_base + 1)].width = 12  # Rank
            ws.column_dimensions[chr(65 + col_base + 2)].width = 55  # Awards

        # Enable text wrapping for better readability
        for i in range(1, years_to_fetch + 1):
            col_base = 4 + (i - 1) * 3
            for j in range(2, len(teams) + 2):
                ws.cell(row=j, column=col_base).alignment = Alignment(wrapText=True)

        # Adjust text alignment for better readability
        for i in range(1, len(teams) + 2):
            ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')
        for i in range(1, years_to_fetch + 1):
            col_base = 2 + (i - 1) * 3
            for j in range(1, len(teams) + 2):
                ws.cell(row=j, column=col_base).alignment = Alignment(horizontal='center')
                ws.cell(row=j, column=col_base + 1).alignment = Alignment(horizontal='center')
        
        wb.save(fullname)
        
        # Fetch data with progress tracking
        print(f"\nFetching data for {len(teams)} teams...")

        all_data = []
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all tasks
            futures = {
                executor.submit(
                    self.fetch_team_data, 
                    team, 
                    start_year, 
                    event_year
                ): team 
                for team in teams
            }
            
            # Process completed tasks
            completed = 0
            for future in as_completed(futures):
                team = futures[future]
                completed += 1
                
                try:
                    data = future.result()
                    
                    # Write to Excel
                    all_data.append(data)
                    
                    # Update progress
                    progress = (completed / len(teams)) * 100
                    print(f"Progress: {completed}/{len(teams)} teams ({progress:.1f}%)", end='\r')
                    
                except Exception as e:
                    logger.error(f"Failed to fetch data for team {team}: {e}")
        
        # Sort all data by team number
        all_data.sort(key=lambda x: x[0])

        # Export data to Excel
        print(f"\nExporting data to {filename}...")

        wb = load_workbook(fullname)
        ws = wb.active

        # Write data to cell separately since the cells are already created (thus appending will result in wrong position)
        for row in range(1, len(all_data) + 1):
            for col in range(1, len(all_data[row - 1]) + 1):
                ws.cell(row + 1, col).value = all_data[row - 1][col - 1]

        wb.save(fullname)

        print(f"✓ Data export complete: {filename}")


def get_user_input() -> tuple:
    """Get and validate user input"""
    while True:
        try:
            print("\n=== FRC Event Data Fetcher ===")
            event_year = int(input("Event year: "))
            if event_year < 1992 or event_year > 2025:
                print("Please enter a valid year (1992-2025)")
                continue
            
            event_code = input("Event code (e.g., txhou, casj): ").strip().lower()
            if not event_code:
                print("Event code cannot be empty")
                continue
            
            self_team_number = int(input("Your team number (optional, press Enter to skip): ") or "0")
            
            years_to_fetch = int(input(f"Years of history to fetch (1-{min(5, event_year - 1991)}): "))
            if years_to_fetch < 1 or years_to_fetch > min(5, event_year - 1991):
                print(f"Please enter a value between 1 and {min(5, event_year - 1991)}")
                continue
            
            return event_year, event_code, self_team_number, years_to_fetch
            
        except ValueError:
            print("Please enter valid numbers")
        except KeyboardInterrupt:
            print("\n\nExiting...")
            sys.exit(0)


def main():
    """Main execution function"""
    try:
        # Get user input
        event_year, event_code, self_team_number, years_to_fetch = get_user_input()
        
        # Initialize fetcher
        config = Config()
        fetcher = FRCDataFetcher(config)
        
        # Construct full event key
        event_key = f"{event_year}{event_code}"
        
        # Fetch teams
        print(f"\nFetching teams for {event_key}...")
        teams = fetcher.get_event_teams(event_key)
        print(f"Found {len(teams)} teams: {teams[:5]}{'...' if len(teams) > 5 else ''}")
        
        # Highlight self team if provided
        if self_team_number in teams:
            print(f"✓ Your team ({self_team_number}) is registered for this event")
        elif self_team_number > 0:
            print(f"✗ Your team ({self_team_number}) is not registered for this event")
        
        # Export data
        fetcher.export_to_excel(event_year, event_code, teams, years_to_fetch)
        
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        print(f"\n❌ Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()