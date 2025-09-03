#!/usr/bin/env python3
"""
FRC Event Data Fetcher Web Server
Web interface for the FRC Event Data Fetcher with download capabilities
"""

import os
import threading
import queue
import time
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import secrets

# Import the main fetcher module
from frc_data_fetcher import FRCDataFetcher, Config

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)
CORS(app)

# Global variables for progress tracking
progress_queue = queue.Queue()
current_tasks = {}

# HTML template with modern UI
template_name = 'web_server.html'

class FetchTask(threading.Thread):
    """Background task for fetching FRC data"""
    
    def __init__(self, task_id, params):
        super().__init__()
        self.task_id = task_id
        self.params = params
        self.progress = 0
        self.status = 'running'
        self.message = 'Initializing...'
        self.detail = ''
        self.filename = None
        
    def run(self):
        try:
            # Initialize fetcher
            config = Config()
            fetcher = FRCDataFetcher(config)
            
            event_year = self.params['event_year']
            event_codes = self.params['event_codes']
            team_number = self.params['team_number']
            years_to_fetch = self.params['years_to_fetch']
            deep_search = self.params['deep_search']
            deep_search_years = self.params['deep_search_years']
            
            total_events = len(event_codes)
            
            for idx, event_code in enumerate(event_codes):
                base_progress = (idx / total_events) * 100
                
                # Construct event key
                event_key = f"{event_year}{event_code}"
                
                # Update progress
                self.message = f"Processing event {event_code} ({idx + 1}/{total_events})"
                self.progress = base_progress
                
                # Fetch teams
                teams = []
                if deep_search:
                    self.detail = f"Deep searching {event_code} across {deep_search_years} years..."
                    for year in range(event_year - deep_search_years + 1, event_year + 1):
                        year_event_key = f"{year}{event_code}"
                        year_teams = fetcher.get_event_teams(year_event_key)
                        teams.extend(year_teams)
                        sub_progress = ((year - (event_year - deep_search_years + 1)) / deep_search_years) * (100 / total_events)
                        self.progress = base_progress + sub_progress * 0.3
                    teams = list(set(teams))  # Remove duplicates
                else:
                    self.detail = f"Fetching teams for {event_key}..."
                    teams = fetcher.get_event_teams(event_key)
                
                if not teams:
                    self.detail = f"No teams found for {event_key}"
                    continue
                
                # Create custom export method with progress callback
                self.detail = f"Fetching data for {len(teams)} teams..."
                self.export_with_progress(fetcher, event_year, event_code, teams, 
                                         years_to_fetch, deep_search, base_progress, 
                                         100 / total_events)
                
                self.filename = f"{event_year}{event_code}{'_deep' if deep_search else ''}.xlsx"
            
            self.status = 'completed'
            self.progress = 100
            self.message = 'Data fetch completed successfully!'
            
        except Exception as e:
            self.status = 'error'
            self.message = str(e)
    
    def export_with_progress(self, fetcher, event_year, event_code, teams, 
                            years_to_fetch, deep_search, base_progress, progress_range):
        """Modified export method with progress updates"""
        import concurrent.futures
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment
        
        foldername = "output"
        filename = f"{event_year}{event_code}{'_deep' if deep_search else ''}.xlsx"
        fullname = os.path.join(foldername, filename)
        
        # Create output folder
        if not os.path.exists(foldername):
            os.makedirs(foldername)
        
        # Remove existing file
        if os.path.exists(fullname):
            os.remove(fullname)
        
        # Create workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.title = f"{event_year} {event_code} Data"
        
        # Create headers
        headers = ['Team']
        start_year = event_year - years_to_fetch + 1
        for year in range(start_year, event_year + 1):
            headers.extend([f'{year} EPA', f'{year} Rank', f'{year} Awards'])
        headers.extend(['Wins', 'Finalists', 'Impact', 'EI'])
        ws.append(headers)
        
        # Adjust column widths for better readability
        ws.column_dimensions['A'].width = 10 # Team number
        for i in range(1, years_to_fetch + 1):
            col_base = 1 + (i - 1) * 3
            ws.column_dimensions[chr(65 + col_base)].width = 12  # EPA
            ws.column_dimensions[chr(65 + col_base + 1)].width = 12  # Rank
            ws.column_dimensions[chr(65 + col_base + 2)].width = 60  # Awards
        for i in range (1 + years_to_fetch * 3, 1 + years_to_fetch + 5):
            ws.column_dimensions[chr(65 + i)].width = 12 # Summary

        # Enable text wrapping for better readability
        for i in range(4, years_to_fetch * 3 + 4, 3):
            for j in range(1, len(teams) + 2):
                ws.cell(row=j, column=i).alignment = Alignment(wrap_text=True) # Awards

        # Adjust text alignment for better readability
        for i in [col for col in range(1, 1 + years_to_fetch * 3 + 5) if col not in [award_col for award_col in range(4, years_to_fetch * 3 + 4, 3)]]:
            for j in range(1, len(teams) + 2):
                ws.cell(row=j, column=i).alignment = Alignment(horizontal='center') # Team number, EPA, Rank, Summary
        
        wb.save(fullname)
        
        # Fetch data with progress tracking
        all_data = []
        completed = 0
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            futures = {
                executor.submit(fetcher.fetch_team_data, team, start_year, event_year): team 
                for team in teams
            }
            
            for future in concurrent.futures.as_completed(futures):
                team = futures[future]
                completed += 1
                
                try:
                    data = future.result()
                    all_data.append(data)
                    
                    # Update progress
                    team_progress = (completed / len(teams)) * progress_range * 0.7
                    self.progress = base_progress + progress_range * 0.3 + team_progress
                    self.detail = f"Processed {completed}/{len(teams)} teams"
                    
                except Exception as e:
                    print(f"Failed to fetch data for team {team}: {e}")
        
        # Sort and write data
        all_data.sort(key=lambda x: x[0])
        
        wb = load_workbook(fullname)
        ws = wb.active
        
        for row in range(1, len(all_data) + 1):
            for col in range(1, len(all_data[row - 1]) + 1):
                ws.cell(row + 1, col).value = all_data[row - 1][col - 1]
        
        wb.save(fullname)


@app.route('/')
def index():
    """Serve the main page"""
    return render_template(template_name)


@app.route('/api/fetch', methods=['POST'])
def start_fetch():
    """Start a new fetch task"""
    data = request.json
    task_id = str(int(time.time() * 1000))
    
    # Create and start background task
    task = FetchTask(task_id, data)
    current_tasks[task_id] = task
    task.start()
    
    return jsonify({'task_id': task_id})


@app.route('/api/progress/<task_id>')
def get_progress(task_id):
    """Get progress of a fetch task"""
    if task_id in current_tasks:
        task = current_tasks[task_id]
        return jsonify({
            'status': task.status,
            'progress': task.progress,
            'message': task.message,
            'detail': task.detail,
            'filename': task.filename
        })
    return jsonify({'status': 'error', 'message': 'Task not found'})


@app.route('/api/files')
def list_files():
    """List all available Excel files"""
    output_dir = Path('output')
    if not output_dir.exists():
        return jsonify([])
    
    files = []
    for file in output_dir.glob('*.xlsx'):
        files.append({
            'name': file.name,
            'size': file.stat().st_size,
            'modified': file.stat().st_mtime * 1000  # Convert to milliseconds
        })
    
    # Sort by modified time (newest first)
    files.sort(key=lambda x: x['modified'], reverse=True)
    return jsonify(files)


@app.route('/download/<filename>')
def download_file(filename):
    """Download a specific file"""
    file_path = Path('output') / filename
    if file_path.exists() and file_path.suffix == '.xlsx':
        return send_file(file_path, as_attachment=True, download_name=filename)
    return "File not found", 404

@app.route('/delete/<filename>', methods=['DELETE'])
def delete_file(filename):
    """Delete a specific file"""
    file_path = Path('output') / filename
    if file_path.exists() and file_path.suffix == '.xlsx':
        file_path.unlink()
        return "File deleted", 200
    return "File not found", 404


if __name__ == '__main__':
    try:
        # Ensure output directory exists
        Path('output').mkdir(exist_ok=True)
        
        # Run the Flask app
        app.run(debug=False, host='0.0.0.0', port=7130)
    except KeyboardInterrupt:
        print("Shutting down...")
    except Exception as e:
        print(f"Error: {e}")